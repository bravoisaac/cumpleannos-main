import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email import encoders
import os
from datetime import date
from openpyxl import Workbook, load_workbook
from fpdf import FPDF
import fitz

print(fitz.__doc__)

#TODO cambiar correo de origen subir main.py usar excel subido a junji:~

def leer_Excel(filename):
    print("leer excel")
    wb = load_workbook(filename)
    print(wb.sheetnames)
    ws = wb['OFICINA MARZO 2025']
    Leer_Hoja(ws)
    ws = wb['JARDIN MARZO 2025']
    Leer_Hoja(ws)

    #for row in range (1,11):
        #for col in range(0,4):
            #char = chr(65 + col)


def Leer_Hoja(ws):
    today = date.today()
    today = str(today).split(" ")[0]
    print("today")
    print(today)
    year = today.split("-")[0]
    todayGroup = str(today).split("-") 
    #print(todayGroup)
    print(ws)
    for row in range(2, ws.max_row): #empieza en 2 porque nesesita
        birthday = ws['M' + str(row)].value #ws['m' + str(row)].value 
        print(str(row) + "birthday")
        print(birthday)
        birthday = str(birthday).split(" ")[0]
        birthdayGroup = birthday.split("-")
        #print(birthdayGroup)
        if(birthdayGroup[1] == todayGroup[1] and birthdayGroup[2] == todayGroup[2]):
            print("HAPPY BIRTHDAY")
            nombre = ws['f' + str(row)].value
            nombre = nombre.split(" ")[0]
            paterno = ws['D' + str(row)].value.strip()
            materno = ws['E' + str(row)].value.strip()
            #nombrepdf = 'cumple_' + str(row) + ".pdf"
            correo = ws['T' + str(row)].value.strip()
            nombrepdf = "cumple.pdf"

            print(nombre + "|")
            print(paterno + "|")
            print(materno + "|")
            print(nombrepdf)

            crear_pdf(nombre, paterno, materno, birthdayGroup[2], 
                    birthdayGroup[1], year, nombrepdf)
            
            nombreImagen = crear_imagen(nombrepdf)
            enviar_correo(nombreImagen, correo, nombre, paterno, materno)
            #Enviar copia a Marti
            enviar_correo(nombreImagen, 'mferrerol@junji.cl', nombre, paterno, materno)
    pass


def crear_pdf(nombre, apellidoPaterno, apellidoMaterno, dia, mes, año, nombrePdf):
    print("crear pdf")

    class PDF(FPDF):
        def header(self):
            imageUrl = "image_1.png"
            #print(imageUrl)
            self.image(imageUrl, keep_aspect_ratio=True, w=pdf.epw) 
            #font
            self.set_font('times', 'B', 12)
            self.set_text_color(170,170,170)
            #Title
            #self.cell(0, 30, '', border=False, ln=1, align='L')
            #self.cell(0, 5, 'JUNTA NACIONAL', border=False, ln=1, align='L')
            #self.cell(0, 5, 'INFANTILES', border=False, ln=1, align='L')
            #self.cell(0, 5, 'Unidad de Inventarios', border=False, ln=1, align='L')
            #line break
            self.ln(10)
        
        def footer(self):
            imageUrl = "image_2.png"
            self.set_y(-45)
            self.image(imageUrl, keep_aspect_ratio=True, w=pdf.epw)
            self.set_y(-70)
            self.set_font('times', 'B', 12)
            self.set_text_color(170,170,170)
            self.cell(0,0, "", ln=1)
            #self.cell(0,0, "Junta Nacional de Jardines Infantiles-JUNJI", ln=1)
            #self.cell(0,12, "OHiggins Poniente 77 Concepción. 041-2125541", ln=1) #problema con el caracter ’
            #self.cell(0,12, "www.junji.cl", ln=1)
    pdf = PDF('P', 'mm', 'A4')
    pdf.set_margin(0)
    pdf.add_page()

    pdf.set_text_color(9, 100, 175)
    titulo = "cumplel"
    #buscar dia de semana
    print("test dia mes")
    print(dia)
    print(mes)
    if(int(dia) < 10 and dia[0] != '0'):
        dia = "0" + dia
    if(int(mes) < 10 and mes[0] != '0'):
        mes = "0" + mes
    #diccionario

    meses = { 
        "01": "Enero",
        "02": "Febrero",
        "03": "Marzo",
        "04": "Abril",
        "05": "Mayo",
        "06": "Junio",
        "07": "Julio",
        "08": "Agosto",
        "09": "Septiembre",
        "10": "Octubre",
        "11": "Noviembre",
        "12": "Diciembre"
    } 
    texto_principal00 = '¡Feliz Cumpleaños!'
    texto_principal1 = """En este día tan especial, deseamos que tengas un cumpleaños repleto de amor y felicidad. Que tus sueños se cumplan y estés siempre en compañía de la alegría, éxito y momentos llenos de sorpresas maravillosas."""
    texto_principal2 = str(nombre) + " " + str(apellidoPaterno) + " " + str(apellidoMaterno)

    final1 = "¡Te deseamos lo mejor!"
    final2 = "Dirección Regional JUNJI Biobío"

    pdf.set_font('Helvetica', 'B', 30)
    pdf.multi_cell(0, 20, texto_principal00, ln=True, align='C')
    pdf.set_font('Helvetica', 'BU', 40)
    pdf.multi_cell(0, 20, texto_principal2, ln=True, align='C')
    pdf.set_font('Helvetica', 'B', 20)
    pdf.multi_cell(0, 15, texto_principal1, ln=True, align='C')
    pdf.set_font('times', 'B', 30)
    pdf.set_text_color(1, 168, 158)
    pdf.ln()
    pdf.multi_cell(0, 20, final1, ln=True, align='C')
    pdf.multi_cell(0, 20, final2, ln=True, align='C')

    pdf.output(nombrePdf)
    return


def crear_imagen(filename):
    print("crear imagen")

    doc = fitz.open(filename)
    page = doc.load_page(0)
    pix = page.get_pixmap()
    imagename = filename.split(".")
    imagename = imagename[0]
    imagename = imagename + ".png"
    pix.save(imagename)
    return imagename

def enviar_correo(filename, correo, nombre, paterno, materno):
    #correo = "cacastilloc@junji.cl"
    print("enviar_correo")
    remitente = '08junjibiobio@junji.cl'
    destinatario = correo#'mferrerol@junji.cl'
    asunto = 'JUNJI te desea un feliz cumpleaños'
    cuerpo = """
    <html>
        <body>
        <img src='cid:image1'/>
        </body>
    </html>
            """.format(nombre, paterno, materno)
    username = '08junjibiobio@junji.cl'
    password = 'Tijunji2017'

    mensaje = MIMEMultipart()

    mensaje['From'] = remitente
    mensaje['To'] = destinatario
    mensaje['Subject'] = asunto

    with open(filename, "rb") as img_file:
        image = MIMEImage(img_file.read())
        image.add_header("Content-ID", "<image1>")
        mensaje.attach(image)

    mensaje.attach(MIMEText(cuerpo, 'html'))

    texto = mensaje.as_string()
    server_smtp1 = 'smtp.office365.com'
    server_smtp2 = 'smtp-mail.outlook.com'
    server = smtplib.SMTP('smtp.office365.com', port=587)
    server.starttls()
    server.login(username, password)
    #print("before send mail")
    #print(destinatario + "__")
    server.sendmail(remitente, destinatario, texto)
    #print("after send mail")
    server.quit()

#revisar notas
def main():
    print("main")
    #ruta_abs = r'C:\Users\Junji\OneDrive - JUNJI'
    #today = date.today()
    #file1name = "NO_EDITAR.txt"
    #file1 = open(ruta_abs + "\\" + file1name, "r")
    #line = file1.readline()
    #file1.close()
    #lineGroup = line.split("-")
    #if(today.year == int(lineGroup[0]) and today.month == int(lineGroup[1]) and today.day == int(lineGroup[2])):
        #print("ya se reviso el dia de hoy")
        #return
    ruta_abs_excel = "cumple_funcionario_correo.xlsx"
    #file1 = open(ruta_abs + "\\" + file1name, "w")
    #file1.write(str(today) + "\n" + "Se reviso por ultima vez")
    leer_Excel(ruta_abs_excel)
    #crear_pdf('Martín', 'Morales', 'Castro', '21', '12')
    #crear_imagen('cumple.pdf')
    #enviar_correo('cumple.png')
main()
#leer_Excel('cumple_funcionario_correo.xlsx')
#crear_pdf('Marcelo', "Escobar", "Quezada", '1', '1', '1800', "cumple.pdf")
#crear_imagen('cumple.pdf')